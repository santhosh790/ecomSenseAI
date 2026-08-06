import io
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd


def _configure_macos_weasyprint_loader_paths():
    """Ensure Homebrew dylib paths are visible to WeasyPrint on macOS."""
    if sys.platform != "darwin":
        return

    candidate_dirs = []
    for brew_prefix in (Path("/opt/homebrew"), Path("/usr/local")):
        candidate_dirs.append(brew_prefix / "lib")
        for lib_name in (
            "glib",
            "pango",
            "cairo",
            "gdk-pixbuf",
            "harfbuzz",
            "fontconfig",
            "freetype",
            "libffi",
        ):
            candidate_dirs.append(brew_prefix / "opt" / lib_name / "lib")

    existing_dirs = [str(path) for path in candidate_dirs if path.exists()]
    if not existing_dirs:
        return

    current_paths = [
        path for path in os.environ.get("DYLD_FALLBACK_LIBRARY_PATH", "").split(":") if path
    ]
    for path in existing_dirs:
        if path not in current_paths:
            current_paths.append(path)

    os.environ["DYLD_FALLBACK_LIBRARY_PATH"] = ":".join(current_paths)


def _prepare_quantity_fields(df):
    working_df = df.copy()
    working_df["Quantity_Value"] = pd.to_numeric(
        working_df["Quantity"].astype(str).str.extract(r"(\d+\.?\d*)")[0],
        errors="coerce",
    ).fillna(0.0)

    working_df["Unit"] = (
        working_df["Quantity"]
        .astype(str)
        .str.extract(r"\b(KG|KGS|EA|NOS)\b", flags=re.IGNORECASE)[0]
        .str.upper()
        .replace({"KGS": "KG", "NOS": "EA"})
        .fillna("KG")
    )
    return working_df


def consolidate(df):
    if df.empty:
        return df

    working_df = _prepare_quantity_fields(df)
    
    # Extract Date column if it exists (all rows should have same date)
    date_value = None
    if 'Date' in working_df.columns:
        date_value = working_df['Date'].iloc[0] if not working_df.empty else None
    
    # Preserve original order by creating a mapping of first occurrence index
    first_occurrence = {}
    for idx, tamil_name in enumerate(working_df['Tamil Name']):
        if tamil_name not in first_occurrence:
            first_occurrence[tamil_name] = idx
    
    # Add order column based on first occurrence
    working_df['_original_order'] = working_df['Tamil Name'].map(first_occurrence)

    result = (
        working_df.groupby(["Tamil Name", "Unit", "_original_order"])["Quantity_Value"]
        .sum()
        .reset_index()
    )
    
    # Sort by original order, then drop the order column
    result = result.sort_values('_original_order').drop(columns=['_original_order'])
    
    # Note: Date column is NOT added back to result since it's shown in the report header
    # The date_value is only used for extracting the date for the header in export functions

    result.rename(columns={"Quantity_Value": "Total Quantity"}, inplace=True)
    return result


def consolidate_with_client_columns(df):
    if df.empty:
        return df

    if "Client Name" not in df.columns:
        return consolidate(df)

    working_df = _prepare_quantity_fields(df)
    working_df["Client Name"] = working_df["Client Name"].astype(str).str.strip()
    working_df["Client Name"] = working_df["Client Name"].replace({"": "Unknown Client"})
    
    # Extract Date column if it exists (all rows should have same date)
    date_value = None
    if 'Date' in working_df.columns:
        date_value = working_df['Date'].iloc[0] if not working_df.empty else None
    
    # Preserve original order by creating a mapping of first occurrence index
    first_occurrence = {}
    for idx, tamil_name in enumerate(working_df['Tamil Name']):
        if tamil_name not in first_occurrence:
            first_occurrence[tamil_name] = idx
    
    # Add order column based on first occurrence
    working_df['_original_order'] = working_df['Tamil Name'].map(first_occurrence)

    pivot_df = (
        working_df.pivot_table(
            index=["Tamil Name", "Unit", "_original_order"],
            columns="Client Name",
            values="Quantity_Value",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reset_index()
    )
    
    # Sort by original order, then drop the order column
    pivot_df = pivot_df.sort_values('_original_order').drop(columns=['_original_order'])

    client_cols = [col for col in pivot_df.columns if col not in ["Tamil Name", "Unit"]]
    if client_cols:
        pivot_df["Total Quantity"] = pivot_df[client_cols].sum(axis=1)
    else:
        pivot_df["Total Quantity"] = 0.0

    # Build result columns
    result_cols = ["Tamil Name", *client_cols, "Total Quantity", "Unit"]
    
    # Note: Date column is NOT added back to result since it's shown in the report header
    # The date_value is only used for extracting the date for the header in export functions

    return pivot_df[result_cols]


def _categorize_item(name):
    """
    Categorize item as 'fruit' or 'vegetable' based on English name.
    Extract English name from "தமிழ் (ENGLISH)" format.
    """
    if not name:
        return "vegetable"
    
    # Extract English name from Tamil Name format
    english_match = re.search(r'\(([^)]+)\)$', str(name))
    english_name = english_match.group(1).upper() if english_match else str(name).upper()
    
    # List of common fruits
    fruits = {
        'APPLE', 'AVOCADO', 'BANANA', 'BLACKBERRY', 'BLUE BERRY', 'CHERRY', 'CHIKKU',
        'CUSTARD APPLE', 'DATES', 'DRAGON FRUIT', 'FIG', 'GRAPE', 'GRAPES', 'GUAVA',
        'JACK FRUIT', 'JACKFRUIT', 'JAVA PLUM', 'JAMUN', 'KIWI', 'LEMON', 'LIME',
        'LITCHI', 'LYCHEE', 'MANGO', 'MELON', 'MOSSAMBI', 'MUSK MELON', 'ORANGE',
        'PAPAYA', 'PASSION FRUIT', 'PEACH', 'PEAR', 'PEARS', 'PINE APPLE', 'PINEAPPLE',
        'PLUM', 'PLUMS', 'POMEGRANATE', 'RAMBUTAN', 'RASPBERRIES', 'SAPOTA', 'STRAW BERRY',
        'STRAWBERRY', 'SWEET LIME', 'WATER MELON', 'WATERMELON', 'WOOD APPLE',
        'GRAPE FRUIT', 'CASHEW APPLE', 'BERRY AUSTRALIAN'
    }
    
    # Check if any fruit keyword is in the name
    for fruit in fruits:
        if fruit in english_name:
            return "fruit"
    
    return "vegetable"


def _sort_items_by_category(df, name_column='Tamil Name'):
    """
    Sort items with vegetables first, then fruits, each group alphabetically.
    """
    if df.empty or name_column not in df.columns:
        return df
    
    df_sorted = df.copy()
    
    # Extract English name for sorting
    df_sorted['_english_name'] = df_sorted[name_column].astype(str).str.extract(r'\(([^)]+)\)$')[0].fillna(df_sorted[name_column])
    
    # Categorize each item
    df_sorted['_category'] = df_sorted[name_column].apply(_categorize_item)
    
    # Sort: vegetables (0) before fruits (1), then alphabetically by English name
    df_sorted['_category_order'] = df_sorted['_category'].map({'vegetable': 0, 'fruit': 1})
    df_sorted = df_sorted.sort_values(by=['_category_order', '_english_name'], ascending=[True, True])
    
    # Drop temporary columns
    df_sorted = df_sorted.drop(columns=['_english_name', '_category', '_category_order'])
    
    return df_sorted


def export_excel(
    df,
    logo_path="",
    header_text="PKS Fresh",
    above_list_text="காய்கறி பட்டியல்",
    footer_text="",
    client_name="",
    order_date=None,
):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    # Extract date from dataframe if available, otherwise use today
    if "Date" in df.columns and not df.empty:
        date_val = df["Date"].iloc[0]
        try:
            # Parse ISO format (YYYY-MM-DD) and convert to DD-MM-YYYY
            date_obj = datetime.strptime(str(date_val), "%Y-%m-%d")
            date_str = date_obj.strftime("%d-%m-%Y")
        except:
            date_str = date.today().strftime("%d-%m-%Y")
    else:
        date_str = date.today().strftime("%d-%m-%Y")
    if order_date:
        try:
            date_obj = datetime.strptime(str(order_date), "%Y-%m-%d")
            date_str = date_obj.strftime("%d-%m-%Y")
        except:
            pass
    tamil_font_name = "Nirmala UI"
    export_df = df.copy()
    
    # Preserve original extraction order (no sorting)
    
    # Rename columns to Tamil
    column_rename_map = {
        'Tamil Name': 'காய்கறி பெயர்',
        'Total Quantity': 'மொத்த அளவு',
        'Unit': 'அலகு'
    }
    export_df = export_df.rename(columns=column_rename_map)

    if " " not in export_df.columns:
        export_df[" "] = ""

    client_text = str(client_name or "").strip()
    date_line_text = f"தேதி: {date_str}"
    if client_text:
        date_line_text = f"வாடிக்கையாளர்: {client_text}    |    {date_line_text}"
    if str(above_list_text or "").strip():
        date_line_text = f"{date_line_text}    |    {str(above_list_text)}"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Vegetables", startrow=6)

        ws = writer.sheets["Vegetables"]

        ws["A1"] = str(header_text or "")
        ws["A3"] = date_line_text

        ws["A1"].font = Font(size=18, bold=True)
        ws["A3"].font = Font(name=tamil_font_name, size=13)

        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A3"].alignment = Alignment(horizontal="left")

        ws.column_dimensions["A"].width = 36
        ws.row_dimensions[3].height = 24

        header_row = 7
        for col_idx in range(1, len(export_df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}{header_row}"].font = Font(size=12, bold=True)
            if col_idx > 1 and ws.column_dimensions[col_letter].width is None:
                ws.column_dimensions[col_letter].width = 14

        for row_idx in range(header_row + 1, header_row + 1 + len(export_df)):
            ws[f"A{row_idx}"].font = Font(name=tamil_font_name, size=14)
            ws[f"A{row_idx}"].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row_idx].height = 24

        if str(footer_text or "").strip():
            footer_row = header_row + 2 + len(export_df)
            ws[f"A{footer_row}"] = str(footer_text)
            ws[f"A{footer_row}"].font = Font(name=tamil_font_name, size=12, bold=True)
            ws[f"A{footer_row}"].alignment = Alignment(horizontal="left", wrap_text=True)

        if logo_path and Path(logo_path).exists():
            logo_img = XLImage(str(logo_path))
            logo_img.height = 85
            logo_img.width = 150
            ws.add_image(logo_img, "D1")

    return output.getvalue()


def export_pdf(
    df,
    logo_data_uri="",
    header_text="PKS Fresh",
    above_list_text="காய்கறி பட்டியல்",
    footer_text="",
    client_name="",
    order_date=None,
):
    _configure_macos_weasyprint_loader_paths()
    from weasyprint import HTML

    # Extract date from dataframe if available, otherwise use today
    if "Date" in df.columns and not df.empty:
        date_val = df["Date"].iloc[0]
        try:
            # Parse ISO format (YYYY-MM-DD) and convert to DD-MM-YYYY
            date_obj = datetime.strptime(str(date_val), "%Y-%m-%d")
            date_str = date_obj.strftime("%d-%m-%Y")
        except:
            date_str = date.today().strftime("%d-%m-%Y")
    else:
        date_str = date.today().strftime("%d-%m-%Y")
    if order_date:
        try:
            date_obj = datetime.strptime(str(order_date), "%Y-%m-%d")
            date_str = date_obj.strftime("%d-%m-%Y")
        except:
            pass
    client_text = str(client_name or "").strip()
    
    # Preserve original extraction order (no sorting)
    df_sorted = df.copy()
    df = df_sorted

    client_cols = [
        col
        for col in df.columns
        if col not in ["Tamil Name", "Total Quantity", "Unit"]
    ]

    rows_html = ""
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        bg = "#f2f2f2" if i % 2 == 0 else "#ffffff"
        tamil = str(row.get("Tamil Name", ""))
        total_qty = row.get("Total Quantity", "")
        unit = str(row.get("Unit", "KG"))

        try:
            qty_val = float(total_qty)
            qty_str = f"{qty_val:.0f}" if qty_val == int(qty_val) else f"{qty_val:.2f}"
        except (ValueError, TypeError):
            qty_str = str(total_qty)

        client_cells = ""
        for col in client_cols:
            value = row.get(col, "")
            try:
                value_num = float(value)
                value_text = f"{value_num:.0f}" if value_num == int(value_num) else f"{value_num:.2f}"
            except (ValueError, TypeError):
                value_text = str(value)
            client_cells += f'<td class="num">{value_text}</td>'

        rows_html += (
            f'<tr style="background:{bg}">'
            f"<td>{i}</td>"
            f"<td><b>{tamil}</b></td>"
            f"{client_cells}"
            f'<td class="num"><b>{qty_str}</b></td>'
            f'<td class="num"><b>{unit}</b></td>'
            "</tr>"
        )

    client_header_html = "".join([f'<th class="num">{col}</th>' for col in client_cols])

    logo_html = ""
    if logo_data_uri:
        logo_html = f'<img class="brand-logo" src="{logo_data_uri}" alt="Company Logo" />'

    html_content = f"""<!DOCTYPE html>
<html lang=\"ta\">
<head>
<meta charset=\"UTF-8\">
<style>
  @font-face {{
    font-family: 'TamilFont';
        src: local('Noto Sans Tamil'), local('Lohit Tamil'), local('Tamil Sangam MN'), local('Tamil MN');
  }}
  @page {{
    margin: 15mm;
  }}
  body {{
        font-family: 'Noto Sans Tamil', 'Lohit Tamil', 'Tamil Sangam MN', 'Tamil MN', serif;
    margin: 0;
    padding: 0;
    color: #111;
  }}
  .header-section {{
    display: flex;
    align-items: center;
    margin-bottom: 10px;
    padding-bottom: 8px;
    border-bottom: 2px solid #2c3e50;
  }}
  .header-left {{
    display: flex;
    align-items: center;
    flex: 1;
  }}
  .brand-logo {{
    height: 50px;
    width: auto;
    margin-right: 15px;
  }}
  .header-text {{
    flex: 1;
  }}
  .company-name {{
    font-size: 20px;
    font-weight: bold;
    margin: 0;
    line-height: 1.2;
  }}
  .subtitle {{
    font-size: 14px;
    color: #555;
    margin: 2px 0 0 0;
  }}
  .date-line {{
    font-size: 12px;
    color: #444;
    margin-bottom: 8px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
  }}
  thead tr {{
    background: #2c3e50;
    color: #fff;
  }}
  th {{
    padding: 6px 10px;
    font-size: 15px;
    text-align: left;
  }}
  th.num {{ text-align: right; }}
  td {{
    padding: 5px 10px;
    font-size: 14px;
    border-bottom: 1px solid #ddd;
  }}
  td.num {{ text-align: right; }}
  tfoot td {{
    font-weight: bold;
    border-top: 2px solid #2c3e50;
    padding: 6px 10px;
    font-size: 11px;
  }}
    .footer-note {{
        margin-top: 14px;
        font-size: 12px;
        color: #222;
        font-weight: 600;
    }}
</style>
</head>
<body>
    <div class=\"header-section\">
        <div class=\"header-left\">
            {logo_html}
            <div class=\"header-text\">
                <div class=\"company-name\">{str(header_text or '')}</div>
                <div class=\"subtitle\">காய்கறி பட்டியல்</div>
            </div>
        </div>
    </div>
    <div class=\"date-line\">{'வாடிக்கையாளர்: ' + client_text + '  |  ' if client_text else ''}{str(above_list_text or '')}{'  |  ' if str(above_list_text or '').strip() else ''}தேதி: {date_str}</div>
  <table>
    <thead>
      <tr>
        <th>#</th>
        <th>காய்கறி பெயர்</th>
        {client_header_html}
        <th class=\"num\">அளவு</th>
        <th class=\"num\">அலகு</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
    <tfoot>
      <tr>
                <td colspan=\"2\">மொத்தம் {len(df)} பொருட்கள்</td>
                <td class=\"num\" colspan=\"{len(client_cols) + 2}\"></td>
      </tr>
    </tfoot>
  </table>
    <div class=\"footer-note\">{str(footer_text or '')}</div>
</body>
</html>"""

    output = io.BytesIO()
    HTML(string=html_content).write_pdf(output)
    return output.getvalue()


def export_delivery_challan_excel(
    df,
    invoice_no="20689",
    invoice_date="",
    po_date="29-07-2026",
    po_delivery_date="30/07/2026",
    vehicle_number="TN23C P8348",
    bill_to_name="RASSENSE PRIVATE LIMITED",
    bill_to_address="No. 15,16,17 2nd Floor,\nVision Towers, Yogam Garden,\nBrindavan Nagar,\nValasaravakkam,\nChennai - 600 087",
    ship_to_name="RASSENSE PRIVATE LIMITED",
    ship_to_address="M/S LARSEN & TOUBRO LIMITED,\nVALVES LIMITED,\nNEXT TO SOSVMV University,\nEnathur Village, Kanchipuram - 631 561",
    payment_mode="Credit",
    company_name="PKS FRESH",
    company_address="1971, M.P SARATHI MANSION,\nNETHAJI MARKET,\nCHENNAI,\n652004",
    phone="9790139595",
    email="pksfresh1@gmail.com",
    logo_path="",
    dl_no="",
    invoice_amount="11319.0",
):
    """Export individual order as delivery challan in Excel format."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    if not invoice_date:
        # Extract date from dataframe if available, otherwise use today
        if "Date" in df.columns and not df.empty:
            date_val = df["Date"].iloc[0]
            try:
                # Parse ISO format (YYYY-MM-DD) and convert to DD-MM-YYYY
                date_obj = datetime.strptime(str(date_val), "%Y-%m-%d")
                invoice_date = date_obj.strftime("%d-%m-%Y")
            except:
                invoice_date = date.today().strftime("%d-%m-%Y")
        else:
            invoice_date = date.today().strftime("%d-%m-%Y")
    
    # Preserve original extraction order (no sorting)
    df_sorted = df.copy()
    df = df_sorted

    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Challan"

    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10

    # Header styles
    header_font = Font(size=16, bold=True)
    subheader_font = Font(size=11, bold=True)
    normal_font = Font(size=10)
    tamil_font = Font(size=16, name="Nirmala UI")  # Increased from 10 to 16
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title: Delivery Challan (centered)
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Delivery Challan'
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Company Name and Details
    current_row = 3
    ws[f'A{current_row}'] = company_name.upper()
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    ws.merge_cells(f'A{current_row}:B{current_row}')
    
    current_row += 1
    ws[f'A{current_row}'] = company_address.replace('\n', ', ')
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'].font = Font(size=9)
    
    current_row += 1
    ws[f'A{current_row}'] = f"Phone no.: {phone}"
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'].font = Font(size=9)
    
    current_row += 1
    ws[f'A{current_row}'] = f"Email: {email}"
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'].font = Font(size=9)

    # Invoice details box (right side)
    invoice_start_row = 3
    ws[f'C{invoice_start_row}'] = 'Invoice No.'
    ws[f'D{invoice_start_row}'] = invoice_no
    ws[f'C{invoice_start_row}'].border = border_thin
    ws[f'D{invoice_start_row}'].border = border_thin
    ws[f'C{invoice_start_row}'].font = subheader_font

    invoice_start_row += 1
    ws[f'C{invoice_start_row}'] = 'Date'
    ws[f'D{invoice_start_row}'] = invoice_date
    ws[f'C{invoice_start_row}'].border = border_thin
    ws[f'D{invoice_start_row}'].border = border_thin
    
    invoice_start_row += 1
    ws[f'C{invoice_start_row}'] = 'PO date'
    ws[f'D{invoice_start_row}'] = po_date
    ws[f'C{invoice_start_row}'].border = border_thin
    ws[f'D{invoice_start_row}'].border = border_thin
    
    invoice_start_row += 1
    ws[f'C{invoice_start_row}'] = 'PO Delivery Date'
    ws[f'D{invoice_start_row}'] = po_delivery_date
    ws[f'C{invoice_start_row}'].border = border_thin
    ws[f'D{invoice_start_row}'].border = border_thin
    
    invoice_start_row += 1
    ws[f'C{invoice_start_row}'] = 'Transportation : PKS'
    ws.merge_cells(f'C{invoice_start_row}:D{invoice_start_row}')
    ws[f'C{invoice_start_row}'].border = border_thin
    
    invoice_start_row += 1
    ws[f'C{invoice_start_row}'] = 'Vehicle Number:'
    ws[f'D{invoice_start_row}'] = vehicle_number
    ws[f'C{invoice_start_row}'].border = border_thin
    ws[f'D{invoice_start_row}'].border = border_thin
    
    invoice_start_row += 1
    ws[f'C{invoice_start_row}'] = 'DL No.'
    ws[f'D{invoice_start_row}'] = dl_no
    ws[f'C{invoice_start_row}'].border = border_thin
    ws[f'D{invoice_start_row}'].border = border_thin

    # Bill To and Ship To sections
    current_row = 9
    ws[f'A{current_row}'] = 'Bill To'
    ws[f'A{current_row}'].font = subheader_font
    ws[f'C{current_row}'] = 'Ship To'
    ws[f'C{current_row}'].font = subheader_font
    
    current_row += 1
    ws[f'A{current_row}'] = bill_to_name
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'].font = Font(size=10, bold=True)
    
    ws[f'C{current_row}'] = ship_to_name
    ws.merge_cells(f'C{current_row}:D{current_row}')
    ws[f'C{current_row}'].font = Font(size=10, bold=True)
    
    # Addresses
    bill_lines = bill_to_address.split('\n')
    ship_lines = ship_to_address.split('\n')
    max_lines = max(len(bill_lines), len(ship_lines))
    
    for i in range(max_lines):
        current_row += 1
        if i < len(bill_lines):
            ws[f'A{current_row}'] = bill_lines[i]
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws[f'A{current_row}'].font = Font(size=9)
        if i < len(ship_lines):
            ws[f'C{current_row}'] = ship_lines[i]
            ws.merge_cells(f'C{current_row}:D{current_row}')
            ws[f'C{current_row}'].font = Font(size=9)

    # Items table header
    current_row += 2
    table_start_row = current_row
    
    ws[f'A{current_row}'] = '#'
    ws[f'B{current_row}'] = 'Item name'
    ws[f'C{current_row}'] = 'Quantity'
    ws[f'D{current_row}'] = 'Unit'
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{current_row}'].font = subheader_font
        ws[f'{col}{current_row}'].border = border_thin
        ws[f'{col}{current_row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Items data
    total_quantity = 0
    for idx, (_, row) in enumerate(df.iterrows(), start=1):
        current_row += 1
        
        # Get item name with Tamil
        source_name = str(row.get('Source Name', ''))
        tamil_name = str(row.get('Tamil Name', ''))
        
        # Format: "English Name (தமிழ்)"
        if tamil_name and '(' in tamil_name:
            tamil_only = tamil_name.split('(')[0].strip()
            item_display = f"{source_name} ({tamil_only})"
        else:
            item_display = source_name
        
        quantity_str = str(row.get('Quantity', ''))
        
        # Parse quantity value
        try:
            qty_match = re.search(r'(\d+\.?\d*)', quantity_str)
            if qty_match:
                qty_val = float(qty_match.group(1))
                total_quantity += qty_val
            else:
                qty_val = ""
        except:
            qty_val = ""
        
        # Parse unit
        unit_match = re.search(r'\b(KG|KGS|EA|NOS|Kg|Nos)\b', quantity_str, re.IGNORECASE)
        unit = unit_match.group(1).upper() if unit_match else 'Kg'
        if unit in ['KGS', 'KG']:
            unit = 'Kg'
        
        ws[f'A{current_row}'] = idx
        ws[f'B{current_row}'] = item_display
        ws[f'C{current_row}'] = qty_val if qty_val else ''
        ws[f'D{current_row}'] = unit
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = border_thin
            ws[f'{col}{current_row}'].font = tamil_font
        
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        ws[f'D{current_row}'].alignment = Alignment(horizontal='center')

    # Total row
    current_row += 1
    ws[f'A{current_row}'] = 'Total'
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].border = border_thin
    
    ws[f'C{current_row}'] = total_quantity
    ws[f'C{current_row}'].font = subheader_font
    ws[f'C{current_row}'].border = border_thin
    ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
    
    ws[f'D{current_row}'].border = border_thin

    # Payment Mode
    current_row += 2
    ws[f'A{current_row}'] = 'Payment Mode'
    ws[f'A{current_row}'].font = subheader_font
    current_row += 1
    ws[f'A{current_row}'] = payment_mode
    ws.merge_cells(f'A{current_row}:D{current_row}')

    # Signature sections
    current_row += 3
    ws[f'A{current_row}'] = 'Received By:'
    ws[f'C{current_row}'] = 'Delivered By:'
    ws[f'A{current_row}'].font = subheader_font
    ws[f'C{current_row}'].font = subheader_font
    
    current_row += 1
    ws[f'A{current_row}'] = 'Name:'
    ws[f'C{current_row}'] = 'Name:'
    
    current_row += 1
    ws[f'A{current_row}'] = 'Comment:'
    ws[f'C{current_row}'] = 'Comment:'
    
    current_row += 1
    ws[f'A{current_row}'] = 'Date:'
    ws[f'C{current_row}'] = 'Date:'
    
    current_row += 1
    ws[f'A{current_row}'] = 'Signature:'
    ws[f'C{current_row}'] = 'Signature:'
    
    ws.merge_cells(f'D{current_row-4}:D{current_row}')
    ws[f'D{current_row-4}'] = f'For: {company_name}'
    ws[f'D{current_row-4}'].font = subheader_font
    ws[f'D{current_row-4}'].alignment = Alignment(horizontal='center', vertical='center')

    # Acknowledgment section
    current_row += 3
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = 'Acknowledgment'
    ws[f'A{current_row}'].font = Font(size=12, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    
    current_row += 1
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = company_name.upper()
    ws[f'A{current_row}'].font = Font(size=11, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    
    # Three-column layout: Invoice To | Invoice Details | Receiver's Seal & Sign
    current_row += 1
    ack_row_start = current_row
    
    # Column 1: Invoice To (A-B)
    ws[f'A{current_row}'] = 'Invoice To:'
    ws[f'A{current_row}'].font = Font(size=9, bold=True)
    ws.merge_cells(f'A{current_row}:B{current_row}')
    
    # Column 2: Invoice Details (C)
    ws[f'C{current_row}'] = 'Invoice Details:'
    ws[f'C{current_row}'].font = Font(size=9, bold=True)
    
    # Column 3: Receiver's Seal & Sign (D)
    ws[f'D{current_row}'] = "Receiver's Seal & Sign"
    ws[f'D{current_row}'].font = Font(size=9, bold=True, italic=True)
    ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
    
    current_row += 1
    
    # Invoice To content
    ws[f'A{current_row}'] = f"{bill_to_name},\n{bill_to_address.replace(chr(10), ', ')}"
    ws.merge_cells(f'A{current_row}:B{current_row+4}')
    ws[f'A{current_row}'].font = Font(size=9)
    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Invoice Details content
    ws[f'C{current_row}'] = f"Invoice No.: {invoice_no}\nInvoice Date: {invoice_date}\nInvoice Amount: {invoice_amount}"
    ws.merge_cells(f'C{current_row}:C{current_row+4}')
    ws[f'C{current_row}'].font = Font(size=9)
    ws[f'C{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Receiver's Seal & Sign space
    ws.merge_cells(f'D{current_row}:D{current_row+4}')
    ws[f'D{current_row}'].border = border_thin
    ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 5

    # Add logo if provided
    if logo_path and Path(logo_path).exists():
        try:
            logo_img = XLImage(str(logo_path))
            logo_img.height = 60
            logo_img.width = 100
            ws.add_image(logo_img, 'A3')
        except:
            pass

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_delivery_challan_pdf(
    df,
    invoice_no="20689",
    invoice_date="",
    po_date="29-07-2026",
    po_delivery_date="30/07/2026",
    vehicle_number="TN23C P8348",
    bill_to_name="RASSENSE PRIVATE LIMITED",
    bill_to_address="No. 15,16,17 2nd Floor,\nVision Towers, Yogam Garden,\nBrindavan Nagar,\nValasaravakkam,\nChennai - 600 087",
    ship_to_name="RASSENSE PRIVATE LIMITED",
    ship_to_address="M/S LARSEN & TOUBRO LIMITED,\nVALVES LIMITED,\nNEXT TO SOSVMV University,\nEnathur Village, Kanchipuram - 631 561",
    payment_mode="Credit",
    company_name="PKS FRESH",
    company_address="1971, M.P SARATHI MANSION,\nNETHAJI MARKET,\nCHENNAI,\n652004",
    phone="9790139595",
    email="pksfresh1@gmail.com",
    logo_data_uri="",
    dl_no="",
    invoice_amount="11319.0",
):
    """Export individual order as delivery challan in PDF format."""
    _configure_macos_weasyprint_loader_paths()
    from weasyprint import HTML

    if not invoice_date:
        # Extract date from dataframe if available, otherwise use today
        if "Date" in df.columns and not df.empty:
            date_val = df["Date"].iloc[0]
            try:
                # Parse ISO format (YYYY-MM-DD) and convert to DD-MM-YYYY
                date_obj = datetime.strptime(str(date_val), "%Y-%m-%d")
                invoice_date = date_obj.strftime("%d-%m-%Y")
            except:
                invoice_date = date.today().strftime("%d-%m-%Y")
        else:
            invoice_date = date.today().strftime("%d-%m-%Y")
    
    # Preserve original extraction order (no sorting)
    df_sorted = df.copy()
    df = df_sorted

    # Build items table rows
    rows_html = ""
    total_quantity = 0
    
    for idx, (_, row) in enumerate(df.iterrows(), start=1):
        bg = "#f9f9f9" if idx % 2 == 0 else "#ffffff"
        
        # Get item name with Tamil
        source_name = str(row.get('Source Name', ''))
        tamil_name = str(row.get('Tamil Name', ''))
        
        # Format: "English Name (தமிழ்)"
        if tamil_name and '(' in tamil_name:
            tamil_only = tamil_name.split('(')[0].strip()
            item_display = f"{source_name} ({tamil_only})"
        else:
            item_display = source_name
        
        quantity_str = str(row.get('Quantity', ''))
        
        # Parse quantity value
        try:
            qty_match = re.search(r'(\d+\.?\d*)', quantity_str)
            if qty_match:
                qty_val = float(qty_match.group(1))
                total_quantity += qty_val
                qty_display = f"{qty_val:.0f}" if qty_val == int(qty_val) else f"{qty_val:.2f}"
            else:
                qty_display = ""
        except:
            qty_display = ""
        
        # Parse unit
        unit_match = re.search(r'\b(KG|KGS|EA|NOS|Kg|Nos)\b', quantity_str, re.IGNORECASE)
        unit = unit_match.group(1) if unit_match else 'Kg'
        if unit.upper() in ['KGS', 'KG']:
            unit = 'Kg'
        
        rows_html += f'''
        <tr style="background: {bg};">
            <td style="text-align: center;">{idx}</td>
            <td><b>{item_display}</b></td>
            <td style="text-align: right;">{qty_display}</td>
            <td style="text-align: center;">{unit}</td>
        </tr>
        '''

    # Total row
    total_display = f"{total_quantity:.1f}"
    
    logo_html = f'<img src="{logo_data_uri}" style="height: 50px;" />' if logo_data_uri else ""

    html_content = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        @page {{
            size: A4;
            margin: 15mm;
        }}
        body {{
            font-family: "Noto Sans Tamil", "Lohit Tamil", Arial, sans-serif;
            font-size: 10pt;
            line-height: 1.4;
            color: #000;
        }}
        .header {{
            text-align: center;
            font-size: 18pt;
            font-weight: bold;
            margin-bottom: 15px;
            border-bottom: 2px solid #000;
            padding-bottom: 5px;
        }}
        .company-info {{
            float: left;
            width: 55%;
        }}
        .company-name {{
            font-size: 14pt;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        .company-details {{
            font-size: 9pt;
            line-height: 1.3;
        }}
        .invoice-box {{
            float: right;
            width: 40%;
            border: 1px solid #000;
        }}
        .invoice-row {{
            display: flex;
            border-bottom: 1px solid #ccc;
        }}
        .invoice-row:last-child {{
            border-bottom: none;
        }}
        .invoice-label {{
            flex: 1;
            padding: 4px 8px;
            font-weight: bold;
            border-right: 1px solid #ccc;
            font-size: 9pt;
        }}
        .invoice-value {{
            flex: 1;
            padding: 4px 8px;
            font-size: 9pt;
        }}
        .clearfix {{
            clear: both;
        }}
        .section-title {{
            font-weight: bold;
            margin-top: 15px;
            margin-bottom: 5px;
            font-size: 10pt;
        }}
        .address-container {{
            display: table;
            width: 100%;
            margin-top: 10px;
        }}
        .address-column {{
            display: table-cell;
            width: 50%;
            vertical-align: top;
            padding-right: 10px;
        }}
        .address-name {{
            font-weight: bold;
            margin-bottom: 3px;
        }}
        .address-text {{
            font-size: 9pt;
            line-height: 1.3;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }}
        th {{
            background: #d3d3d3;
            border: 1px solid #000;
            padding: 6px;
            font-weight: bold;
            text-align: left;
            font-size: 10pt;
        }}
        td {{
            border: 1px solid #ccc;
            padding: 6px;
            font-size: 16px;  /* Increased from 9pt to 16px */
        }}
        .total-row td {{
            font-weight: bold;
            background: #f0f0f0;
        }}
        .payment-section {{
            margin-top: 15px;
        }}
        .signature-section {{
            display: table;
            width: 100%;
            margin-top: 30px;
        }}
        .signature-column {{
            display: table-cell;
            width: 33%;
            vertical-align: top;
        }}
        .signature-label {{
            font-weight: bold;
            margin-bottom: 40px;
        }}
        .acknowledgment {{
            margin-top: 30px;
            border-top: 2px solid #000;
            padding-top: 10px;
        }}
        .ack-title {{
            text-align: center;
            font-size: 12pt;
            font-weight: bold;
            margin-bottom: 10px;
        }}
        .ack-company {{
            font-size: 11pt;
            font-weight: bold;
            margin-bottom: 8px;
            text-align: center;
        }}
        .ack-columns {{
            display: table;
            width: 100%;
            margin-top: 10px;
        }}
        .ack-column {{
            display: table-cell;
            width: 33%;
            vertical-align: top;
            padding: 10px;
            border-right: 1px solid #ccc;
        }}
        .ack-column:last-child {{
            border-right: none;
        }}
        .ack-column-title {{
            font-weight: bold;
            margin-bottom: 8px;
            font-size: 10pt;
        }}
        .ack-details {{
            font-size: 9pt;
            line-height: 1.4;
        }}
        .seal-sign-box {{
            border: 1px solid #ccc;
            height: 80px;
            text-align: center;
            padding-top: 30px;
            font-style: italic;
            font-size: 9pt;
        }}
    </style>
</head>
<body>
    <div class="header">Delivery Challan</div>
    
    <div class="company-info">
        {logo_html}
        <div class="company-name">{company_name}</div>
        <div class="company-details">
            {company_address.replace(chr(10), '<br>')}<br>
            FSSAI 2242159800041<br>
            Phone no.: {phone}<br>
            Email: {email}
        </div>
    </div>
    
    <div class="invoice-box">
        <div class="invoice-row">
            <div class="invoice-label">Invoice No.</div>
            <div class="invoice-value">{invoice_no}</div>
        </div>
        <div class="invoice-row">
            <div class="invoice-label">Date</div>
            <div class="invoice-value">{invoice_date}</div>
        </div>
        <div class="invoice-row">
            <div class="invoice-label">PO date</div>
            <div class="invoice-value">{po_date}</div>
        </div>
        <div class="invoice-row">
            <div class="invoice-label">PO Delivery Date</div>
            <div class="invoice-value">{po_delivery_date}</div>
        </div>
        <div class="invoice-row">
            <div class="invoice-label" style="border-right: none;" colspan="2">Transportation : PKS</div>
        </div>
        <div class="invoice-row">
            <div class="invoice-label">Vehicle Number:</div>
            <div class="invoice-value">{vehicle_number}</div>
        </div>
        <div class="invoice-row">
            <div class="invoice-label">DL No.</div>
            <div class="invoice-value">{dl_no}</div>
        </div>
    </div>
    
    <div class="clearfix"></div>
    
    <div class="address-container">
        <div class="address-column">
            <div class="section-title">Bill To</div>
            <div class="address-name">{bill_to_name}</div>
            <div class="address-text">{bill_to_address.replace(chr(10), '<br>')}</div>
        </div>
        <div class="address-column">
            <div class="section-title">Ship To</div>
            <div class="address-name">{ship_to_name}</div>
            <div class="address-text">{ship_to_address.replace(chr(10), '<br>')}</div>
        </div>
    </div>
    
    <table>
        <thead>
            <tr>
                <th style="width: 5%; text-align: center;">#</th>
                <th style="width: 60%;">Item name</th>
                <th style="width: 20%; text-align: right;">Quantity</th>
                <th style="width: 15%; text-align: center;">Unit</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
            <tr class="total-row">
                <td colspan="2" style="text-align: left;">Total</td>
                <td style="text-align: right;">{total_display}</td>
                <td></td>
            </tr>
        </tbody>
    </table>
    
    <div class="payment-section">
        <div class="section-title">Payment Mode</div>
        <div>{payment_mode}</div>
    </div>
    
    <div class="signature-section">
        <div class="signature-column">
            <div class="signature-label">Received By:</div>
            <div>Name:</div>
            <div>Comment:</div>
            <div>Date:</div>
            <div>Signature:</div>
        </div>
        <div class="signature-column">
            <div class="signature-label">Delivered By:</div>
            <div>Name:</div>
            <div>Comment:</div>
            <div>Date:</div>
            <div>Signature:</div>
        </div>
        <div class="signature-column">
            <div class="signature-label" style="text-align: center; margin-top: 20px;">For: {company_name}</div>
        </div>
    </div>
    
    <div class="acknowledgment">
        <div class="ack-title">Acknowledgment</div>
        <div class="ack-company">{company_name}</div>
        
        <div class="ack-columns">
            <div class="ack-column">
                <div class="ack-column-title">Invoice To:</div>
                <div class="ack-details">
                    <strong>{bill_to_name}</strong><br>
                    {bill_to_address.replace(chr(10), '<br>')}
                </div>
            </div>
            
            <div class="ack-column">
                <div class="ack-column-title">Invoice Details:</div>
                <div class="ack-details">
                    <strong>Invoice No.:</strong> {invoice_no}<br>
                    <strong>Invoice Date:</strong> {invoice_date}<br>
                    <strong>Invoice Amount:</strong> {invoice_amount}
                </div>
            </div>
            
            <div class="ack-column">
                <div class="ack-column-title">Receiver's Seal & Sign:</div>
                <div class="seal-sign-box">
                    (Seal & Signature)
                </div>
            </div>
        </div>
    </div>
</body>
</html>'''

    output = io.BytesIO()
    HTML(string=html_content).write_pdf(output)
    return output.getvalue()
